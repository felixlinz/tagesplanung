from openpyxl import load_workbook
from datetime import datetime, time, timedelta
from tabulate import tabulate


class Day:
    def __init__(self, date):
        self.date = date
        self.touren = []
    
    def add_tour(self, tour):
        self.touren.append(tour)

class Tour:
    def __init__(self, nummer, fahrer, date, beginn = time(hour=7, minute=45)):
        self.date = date
        self.nummer = nummer
        self.fahrer = fahrer
        self.begin = beginn
        self.stadtteil
        

 

def main():
    jahresplan = load_jahresplanung('Jahresplanung.xlsx')
    tag = create_tagesplan(jahresplan)
    for tour in tag.touren:
        print(tour.nummer, tour.fahrer, tour.begin)
    
        

def load_jahresplanung(filename):              
    wb = load_workbook(filename)
    sheet = wb["Jahresplanung"]

    jahresplanung = []

    for row in sheet.iter_cols(values_only=True):
        jahresplanung.append([item for item in row])
        
    return jahresplanung


def create_tagesplan(jahresplan):
    datum = jahresplan[-1][0]

    tag = Day(datum)

    for i, zusteller in enumerate(jahresplan[-1][0:]):
        if jahresplan[0][i] is not None and jahresplan[0][i].isnumeric():
            tour = Tour(jahresplan[0][i], str(zusteller), datum)
            tag.add_tour(tour)
            
    return tag


def write_tagesplan_file(tagesplan):
    # Create a new workbook and select the active worksheet
    tp = Workbook()
    tps = tp.active

    # Define the headers
    headers = ["Bezirk", "Name", "Beginn", "Unterschrift", "Ende", "+0.15"]

    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Save the workbook
    wb.save('new_file.xlsx')
    

    
    
    
    
if __name__== "__main__":
    main()